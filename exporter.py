from __future__ import annotations
from pathlib import Path
import pandas as pd

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _format_header(ws) -> None:
    # Bold + wrap + centered headers, taller row
    header_font = Font(bold=True)
    ws.row_dimensions[1].height = 28  # makes wrapped headers readable

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions  # filter dropdowns like real models


def _smart_autosize(ws, max_width: int = 48, padding: int = 2) -> None:
    """
    Sets column widths based on max(header length, data length) + padding.
    Much better than only looking at data.
    """
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)

        # header length
        header_val = ws.cell(row=1, column=col_idx).value
        header_len = len(str(header_val)) if header_val is not None else 0

        # max data length (scan rows)
        data_len = 0
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            data_len = max(data_len, len(str(v)))

        width = min(max(header_len, data_len) + padding, max_width)
        ws.column_dimensions[col_letter].width = max(width, 10)  # minimum width


def _force_widths(ws) -> None:
    """
    Hard minimum widths for columns that commonly show #### or have long names.
    This is what makes it look good repeatedly.
    """
    forced = {
        "A": 18,  # date column always
    }
    for col_letter, w in forced.items():
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, w)

    # If these headers exist, set comfortable widths (by header name)
    min_by_header = {
        "date": 18,
        "revenue": 16,
        "enterprise_value": 20,
        "ebitda_normalized": 20,
        "interest_expense": 18,
        "accounts_receivable": 22,
        "accounts_payable": 20,
        "ebitda_margin_change_pp": 22,
    }

    headers = [c.value for c in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers) if isinstance(h, str)}

    for header, width in min_by_header.items():
        if header in col_map:
            col_letter = get_column_letter(col_map[header])
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width or 0, width)


def _set_formats(ws) -> None:
    headers = [c.value for c in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(headers) if isinstance(h, str)}

    # Date
    if "date" in col_map:
        col = col_map["date"]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "yyyy-mm-dd"

    # Percent columns
    percent_cols = {"gross_margin", "ebitda_margin", "revenue_yoy", "revenue_cagr", "roic"}
    for h in percent_cols:
        if h in col_map:
            col = col_map[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = "0.00%"

    # Multiples
    for h in {"EV/Revenue", "EV/EBITDA"}:
        if h in col_map:
            col = col_map[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = "0.00x"

    # $mm columns (1 decimal)
    mm_cols = {"revenue", "ebitda_normalized", "fcf", "net_debt", "enterprise_value"}
    for h in mm_cols:
        if h in col_map:
            col = col_map[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).number_format = '#,##0.0" mm"'

    # Percentage points
    if "ebitda_margin_change_pp" in col_map:
        col = col_map["ebitda_margin_change_pp"]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = '0.0" pp"'


def export_to_excel(master: pd.DataFrame, valuation: pd.DataFrame, summary: pd.DataFrame, out_path: str | Path) -> None:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="Master_Model", index=False)
        valuation.to_excel(writer, sheet_name="Valuation_Snapshot", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        wb = writer.book
        for name in ["Master_Model", "Valuation_Snapshot", "Summary"]:
            ws = wb[name]
            _format_header(ws)
            _set_formats(ws)
            _smart_autosize(ws, max_width=48, padding=2)
            _force_widths(ws)
